import copy

from src.library.Common import Database, createdir
from src.library import Common as c
from datetime import datetime
import joblib
from omlt.io.input_bounds import write_input_bounds
import openpyxl
import os
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from src.library.Plot import tempforecast
import random
from sklearn.model_selection import train_test_split
from sklearn.metrics import r2_score
from sklearn import linear_model, preprocessing
from sklearn.preprocessing import StandardScaler
import time
import torch
from torch import nn
import torch.optim as optim
from torch.utils.data import DataLoader, TensorDataset, Dataset
from torch.utils.tensorboard import SummaryWriter  # PyTorch TensorBoard support
from torcheval.metrics.functional import r2_score
from tensorboard import program


class RCmodel(nn.Module):
    def __init__(self, nb_zones, delta_t, target_name):
        super().__init__()
        self.nb_zones = nb_zones
        self.delta_t = delta_t
        self.flatten = nn.Flatten()
        # initialize the efficiency to 3 and 2
        self.heating_efficiency = nn.Parameter(torch.ones(1, nb_zones) * 3)  # Heating efficiency
        self.cooling_efficiency = nn.Parameter(torch.ones(1, nb_zones) * 2)  # Cooling efficiency
        # initialize the R and C to random integers between 5 and 20
        self.R = nn.Parameter(torch.FloatTensor(1, nb_zones).uniform_(5, 21))  # Thermal resistance
        self.C = nn.Parameter(torch.FloatTensor(1, nb_zones).uniform_(5, 21))  # Thermal capacitance
        # self.hour_coeff = nn.Parameter((torch.randn(24, nb_zones) / 1e3).float())
        self.sun_diff_coeff = nn.Parameter(torch.rand(1, nb_zones) / 1e3)
        self.sun_dir_coeff = nn.Parameter(torch.rand(1, nb_zones) / 1e3)
        self.nd_load_coeff = nn.Parameter(torch.rand(1, nb_zones) / 1e3)

    # property decorator to make beta and gamma "callable" as attributes
    # when writing self.beta, the function beta(self) will be called and return the value
    @property
    def beta(self):
        return self.delta_t / self.C

    @property
    def gamma(self):
        return self.delta_t / (self.R * self.C)

    def forward(self, xb):
        t_in, t_amb, p_h, p_c, h, sun_diff, sun_dir, nd_load = (
            xb["t_in"], xb["t_amb"], xb["p_h"], xb["p_c"], xb["hour"], xb["sun_diff"], xb["sun_dir"], xb["nd_load"])
        # dimension check
        if t_in.shape[-1] != self.nb_zones:
            raise ValueError(f"t_in should have the shape (*, {self.nb_zones}) but got {t_in.shape}")
        if t_amb.shape[-1] != 1:
            raise ValueError(f"t_amb should have the shape (*, 1) but got {t_amb.shape}")
        if p_h.shape[-1] != self.nb_zones:
            raise ValueError(f"p_h should have the shape (*, {self.nb_zones}) but got {p_h.shape}")
        if p_c.shape[-1] != self.nb_zones:
            raise ValueError(f"p_c should have the shape (*, {self.nb_zones}) but got {p_c.shape}")
        # forward pass
        prediction = self.flatten(p_h * self.beta * self.heating_efficiency
                                  - p_c * self.beta * self.cooling_efficiency
                                  + (t_amb - t_in) * self.gamma + t_in
                                  # + h @ self.hour_coeff  # h.double()
                                  # + sun_diff * self.sun_diff_coeff + sun_dir * self.sun_diff_coeff
                                  # + self.nd_load_coeff * nd_load
                                  )
        return prediction

class SpatialRCmodel(RCmodel):
    def __init__(self, nb_zones, delta_t, target_name):
        super().__init__(nb_zones, delta_t, target_name)
        # matrix of the heat transfer coeff: alpha
        # initialize to random between 0 and 0.1 (1 on the diagonal if target_name is "Zone Mean Air Temperature(t+1)")
        tmp = torch.FloatTensor(nb_zones, nb_zones).uniform_(0, 0.1)
        self.alpha = nn.Parameter(tmp.masked_fill(torch.diag(torch.ones(nb_zones)).bool(), 1)) if (
                target_name == "Zone Mean Air Temperature(t+1)") else nn.Parameter(tmp)

    def forward(self, xb):
        # necessary to indicate the second dimension of t_in because it is used in the matrix multiplication
        t_in = xb["t_in"].reshape(-1, self.nb_zones)  # reshape to (batch_size, nb_zones)
        RC_prediction = super().forward(xb)
        prediction = self.flatten(t_in @ self.alpha) + RC_prediction - t_in
        return prediction


class RCmodelDataset(Dataset):
    """
    Dataset for the RCmodel problem
    Each sample is a dictionary with the following keys:
    - t_in: the indoor temperature
    - t_amb: the ambient temperature
    - p_h: the heating power
    - p_c: the cooling power
    """

    def __init__(self, data, target_name):
        """

        :param data: a df containing the data
        :param target_name: the zone-agnostic name of target variable. E.g., "Zone Mean Air Temperature(t+1)"
        or Delta Zone Mean Air Temperature(t+1)
        """
        self.data = data
        self.t_in = data.filter(axis=1, like="Zone Mean Air Temperature,")
        self.t_amb = data.filter(axis=1, like="Outdoor Air Drybulb Temperature")
        zonal_sensible_heating_rate = data.filter(axis=1, like="Zone Air System Sensible Heating Rate")
        zonal_sensible_cooling_rate = data.filter(axis=1, like="Zone Air System Sensible Cooling Rate")
        self.p_h = data.filter(axis=1, like="Air System Electricity Energy").where(
            zonal_sensible_heating_rate.values > zonal_sensible_cooling_rate.values, 0).div(1000)
        self.p_c = data.filter(axis=1, like="Air System Electricity Energy").where(
            zonal_sensible_heating_rate.values <= zonal_sensible_cooling_rate.values, 0).div(1000)
        self.hour = pd.DataFrame(data.index.hour, index=data.index)
        self.sun_diffuse = data.loc[:, ["Site Diffuse Solar Radiation Rate per Area,ENVIRONMENT"]].div(1000)
        self.sun_direct = data.loc[:, ["Site Direct Solar Radiation Rate per Area,ENVIRONMENT"]].div(1000)
        self.nd_load = data.filter(axis=1, like="Electricity:Zone:").div(1000)
        self.target = data.filter(axis=1, like=target_name)

    def __len__(self):
        return self.data.shape[0]

    def __getitem__(self, idx):
        def getitem(x):
            return torch.tensor(x.iloc[idx].values, dtype=torch.float32)

        def h_onehotencoding(h):
            a = np.zeros(24)
            a[c.floor(h)] = 1
            return a

        input = dict(zip(["t_in", "t_amb", "p_h", "p_c", "sun_diff", "sun_dir", "nd_load"],
                         [getitem(x) for x in [
                             self.t_in, self.t_amb, self.p_h, self.p_c, self.sun_diffuse, self.sun_direct,
                             self.nd_load]]))
        input["hour"] = h_onehotencoding(getitem(self.hour))
        target = getitem(self.target)
        return input, target


class WrappedDataLoader:
    def __init__(self, dl, func):
        self.dl = dl
        self.func = func

    def __len__(self):
        return len(self.dl)

    def __iter__(self):
        for b in self.dl:
            # batch b is a list containing one dictionary (with keys: "t_in", "t_amb", "p_h" and "p_c")
            # and one tensor (the target)
            # The value to each key is a tensor with first dimension = batch size
            yield (self.func(*b))


def loss_batch(model, loss_func, xb, yb, opt=None):
    yb_hat = model(xb)
    losses = {"mse": nn.MSELoss()(yb_hat, yb), "rmse": torch.sqrt(nn.MSELoss()(yb_hat, yb)),
              "mae": nn.L1Loss()(yb_hat, yb), "mape": torch.mean(torch.abs((yb_hat - yb) / yb)) * 100,
              "r2": r2_score(yb_hat, yb), "std": torch.std(yb_hat - yb), "rstd": torch.std((yb_hat - yb) / yb)}

    if opt is not None:
        losses[loss_func].backward()
        opt.step()
        opt.zero_grad()

    # returns a copy of the tensor loss as a python number
    for k, v in losses.items():
        losses[k] = v.item()

    # get the actual batch size (the last batch may have a different size)
    actual_batch_size = len(xb.get(xb.keys().__iter__().__next__()))
    return losses, actual_batch_size


def fit(hp, model, loss_func, opt, train_dl, val_dl):
    losses_train, losses_val = [], []
    patience, best_epoch = 0, 0
    best_model = copy.deepcopy(model)
    for epoch in range(hp["nb_epochs_max"]):
        model.train()
        # train the model
        losses_train_ep, nums_train = zip(*[loss_batch(model, loss_func, xb, yb, opt) for xb, yb in train_dl])

        def weighted_avg(x, w):
            return np.round(np.sum(np.multiply(x, w)) / np.sum(w), decimals=3)

        # compute the average losses for the epoch across the batches
        losses_train.append(c.agg_itr_of_dict(losses_train_ep, weighted_avg, nums_train))

        model.eval()
        with torch.no_grad():
            losses_val_ep, nums_val = zip(*[loss_batch(model, loss_func, xb, yb) for xb, yb in val_dl])
        losses_val.append(c.agg_itr_of_dict(losses_val_ep, weighted_avg, nums_val))
        print(f"Epoch {epoch}/{hp['nb_epochs_max']}: train_loss = {losses_train[-1][loss_func]}"
              f" - val. {losses_val[-1]}")

        # early stopping
        if losses_val[-1][loss_func] < losses_val[best_epoch][loss_func]:
            best_epoch = epoch
            best_model = copy.deepcopy(model)
            patience = 0
        else:
            patience += 1
            if patience > hp["patience_max"]:
                print(f"Early stopping at epoch {epoch}")
                break

    return losses_train, losses_val, best_epoch, best_model


def rcmodelall(model, data: pd.DataFrame, target_name, hp: dict, model_folderpath: str,
               save_models: bool = False, seed: int = None):
    """
    Train and test the RC model on the given data. The model is saved in the specified folder.
    Args:
        model: RCmodel or SpatialRCmodel
        data:
        target_name:
        hp: hyperparameters
        model_folderpath:
        save_models:
        seed:

    Returns:

    """

    # To make code reproducible
    seed, seed_worker, g = c.make_reproducible(seed)

    # device available
    # IMPORTANT: to debug, force the device to be the cpu (dev = "cpu")
    dev = "cpu"  # "cpu" or (does not work with the current code on other than cpu)  c.get_available_device()

    # prepare the dataloaders
    train_data, val_data = train_test_split(data, test_size=0.2, shuffle=True, random_state=seed)
    train_ds, val_ds = map(lambda x: RCmodelDataset(x, target_name), (train_data, val_data))
    train_dl = DataLoader(train_ds, batch_size=hp["batch_size"], shuffle=True, worker_init_fn=seed_worker, generator=g)
    # double batch size because no backprop needed (so no gradient to store in memory)
    val_dl = DataLoader(val_ds, batch_size=hp["batch_size"] * 2, shuffle=True, worker_init_fn=seed_worker, generator=g)

    # to preprocess the data fed by the dataloader
    def preprocess_data(x, y):
        """
        Used in the WrappedDataLoader class. Preprocess the data before feeding it to the model.
        The data are transformed into tensors and sent to the device.
        :param x: a dictionary containing the input. Keys are "t_in", "t_amb", "p_h" and "p_c".
                Each value is a tensor with first dimension = batch size
        :param y: a tensor containing the target with shape (batch_size, nb_zones)
        :return:
        """
        preprocessed_x = {k: v.to(dev, dtype=torch.float32) for k, v in x.items()}
        preprocessed_y = y.to(dev, dtype=torch.float32)
        return preprocessed_x, preprocessed_y

    train_dl = WrappedDataLoader(train_dl, preprocess_data)
    val_dl = WrappedDataLoader(val_dl, preprocess_data)

    # create the model
    delta_t = data.index.freq.delta.total_seconds() / 3600  # time step between each row of the dataset
    rcmodel = model(nb_zones=hp["nb_zones"], delta_t=delta_t, target_name=target_name)
    rcmodel.to(dev)

    # loss function (mse, rmse, mae, mape, r2)
    loss_func = "mse"
    print(f"Loss function: {loss_func}")

    # optimizer
    opt = optim.Adam(rcmodel.parameters(), lr=hp["lr"])

    # fit the model
    losses_train, losses_val, best_epoch, best_model = fit(hp, rcmodel, loss_func, opt, train_dl, val_dl)

    # Prepare 24h-ahead forecast
    def make_dl_iter_a_sample(xb, yb):
        """
        take the input batch and the target batch. Input batch is a dictionary whose values are tensors.
        Tensors sizes are batch size x nb_zones. The function returns a list of batch size dictionaries
        with the same keys but with values being single values (not tensors).
        So it is easier to iterate over the batch for the 24h forecast.
        :param xb:
        :param yb:
        :return: the batch input which is a list of dictionaries and the target batch
        """
        return [{k: v[i].to(dev, dtype=torch.float32) for k, v in xb.items()} for i in range(len(yb))], yb

    test24h_dl = WrappedDataLoader(DataLoader(RCmodelDataset(data, target_name),
                                              batch_size=24, shuffle=False, worker_init_fn=seed_worker, generator=g,
                                              drop_last=True),
                                   make_dl_iter_a_sample)
    _, metrics24h = metricsover24hours_rc(best_model, target_name, test24h_dl)

    print("RC h_eff: ", best_model.heating_efficiency, "\n RC c_eff: ", best_model.cooling_efficiency)

    if save_models:
        # create the folder for the model
        now = datetime.now().strftime("%Y-%m-%d %Hh%Mm%Ss")
        c.createdir(os.path.join(model_folderpath, now), up=0)

        # save the model to pickle
        joblib.dump(best_model, os.path.join(model_folderpath, now, "RCmodel.pkl"))
        # grab one single batch
        one_batch = next(iter(train_dl))
        # select the input of the first batch and, given that it is a dictionary, specify that it is not a kwarg
        # cf. https://pytorch.org/docs/stable/onnx_torchscript.html#torch.onnx.export
        input_sample = (one_batch[0], {})
        # save the model to onnx
        to_onnx(os.path.join(model_folderpath, now, "RCmodel.onnx"), best_model, input_sample)
        # save to torch model
        torch.save(best_model.state_dict(), os.path.join(model_folderpath, now, "RCmodel.pth"))
        # save the model to the summary
        to_save = {"time": now, "target": target_name, "seed": seed, "lr": hp["lr"],
                   "last_epoch": f"{len(losses_train)}/{hp['nb_epochs_max']}", "best_epoch": best_epoch,
                   "loss_func": loss_func, "losses_train": losses_train[best_epoch][loss_func]}
        to_save.update(losses_val[best_epoch])
        to_save.update(metrics24h)
        modelclassstr = "RC" if model.__name__ == "RCmodel" else "spatialRC"
        c.save_newline_excel(os.path.join(model_folderpath, f"summary_{modelclassstr}.xlsx"), to_save, index_col=None, header=0)


def LRtrainingandtest(db: Database, intercept: bool = True):
    """

    :param db: should be the normalized database
    :return: the trained model, the scaler for the data, and the metrics (dict)
    """
    # poly = preprocessing.PolynomialFeatures(degree=1)
    reg = linear_model.LinearRegression(fit_intercept=intercept)
    db_LR = Database(pd.concat([db.X, db.y_true], axis=1), db.y_true.columns, db.seed)  # a copy of db
    # db_LR.X_train, db_LR.X_test = (poly.fit_transform(db_LR.X_train), poly.fit_transform(db_LR.X_test))
    reg, metrics = c.modeltrainandtest(db_LR, reg)
    print("\nCoefficients:\t", reg.coef_)
    print("Intercept:\t", reg.intercept_)

    return reg, metrics


def LRsave(lreg, metrics, input_variable_names, filepath):
    joblib.dump(lreg, f"{filepath}LR.pkl")
    parameter_names = ["Intercept"] + input_variable_names
    parameter_values = list(lreg.intercept_) + list(lreg.coef_[0, 1:])

    df_parameters = pd.DataFrame(dict(zip(parameter_names, parameter_values)), index=range(1))
    df_metrics = pd.DataFrame(np.round([v for v in metrics.values()], decimals=3), index=metrics.keys(),
                              columns=["Metrics"])

    frames = {'Metrics': df_metrics,
              'Parameters': df_parameters}  # store the df in a dict, where the key is the sheet name
    with pd.ExcelWriter(f"{filepath}LR.xlsx", engine='openpyxl') as writer:  # initialize the writer
        for sheet, frame in frames.items():  # loop through the dict and save each df
            frame.to_excel(writer, sheet_name=sheet)


def PBEpredict(X, pbe, scal=None):
    Tin_pts = X["Zone Mean Air Temperature [C](t-0)"]
    Ta_pts = X["Outdoor Air Drybulb Temperature [C](t-0)"]
    p_heating_pts = X['Heating Coil Electricity Rate [W](t-0)'] / 1000
    p_cooling_pts = X['Cooling Coil Electricity Rate [W](t-0)'] / 1000
    t_in = pbe.predict(Tin_pts, Ta_pts, p_heating_pts, p_cooling_pts)
    return t_in


def LRpredict(X, lreg, scal: dict, order=1):
    """
    Return the prediction of the LR model based on the input X. /!\ The degree of the LR is assumed to be one
    :param X: a numpy array which contains with n° rows = n° samples and n° columns = n° features
    :param lreg: the linear regression model
    :param scal: a dictionary containing the scalers used to scale the input X and output Y
    :param order: the degree of the polynomial model (here, linear regression => 1)
    :return: the predicted value of the output
    """
    scal_x, scal_y = scal.get('input'), scal.get('output')
    X = pd.DataFrame(X.values.reshape((1, -1)), columns=X.index)  # convert X (a pd.Series) into a pd.DataFrame
    tmp = scal_x.transform(X)
    # tmp = preprocessing.PolynomialFeatures(degree=order).fit_transform(X)
    X.iloc[0, :] = tmp  # convert X (a pd.Series) into a pd.DataFrame
    scaled_output = lreg.predict(X)
    return scal_y.inverse_transform(scaled_output)


def RFpredict(X, rfreg, scal: StandardScaler = None):
    """
    Return the prediction of the Random Forest model based on the input X.
    :param X: a pd.Series of size n° features
    :param scal: the scaler used to scale the input X (none for the RF)
    :param rfreg: the random forest or the gradient boosting trees model
    :return: the predicted value of the output
    """
    return rfreg.predict(X.values.reshape((1, -1)))


def NNpredict(X, model, scalers: dict, device='cpu'):
    model.eval()  # set model to evaluation mode
    scal_x, scal_y = scalers.get('input'), scalers.get('output')
    X = pd.DataFrame(dict(zip(X.index, X.values)), index=[0])
    X = scal_x.transform(X)
    X = torch.tensor(X, dtype=torch.double)
    with torch.no_grad():  # disable gradient computation
        model.to('cpu')  # move model to cpu (does not work on mps)
        X.to('cpu')
        tmp = model(X)
    return scal_y.inverse_transform(tmp)


def metricsover24hours_nn(model, target_name, dl, temp_pos_idx: list):
    """
    Compute the metrics over 24h for the NN model. Same function as for the RC model but adapted to the
    dataloader which returns a tensor for each batch and not a dictionary of tensor.
    """
    y_true_l, y_hat_l = [], []
    true_temp_l, forecasted_temp_l = [], []  # these are useful only for Delta Zone Mean Air Temperature(t+1) target name
    model.eval()
    with torch.no_grad():
        for i_b, (xb, yb) in enumerate(dl):
            if i_b % 4 == 0:  # only one day out of 4 is considered to quicken the computation
                forecasted_temp = xb[0][temp_pos_idx]
                for xs, ys in zip(xb, yb):
                    true_t_in = xs[temp_pos_idx]
                    xs[temp_pos_idx] = forecasted_temp
                    y_hat_l.append(model(xs))
                    y_true_l.append(ys)
                    # if we predict directly the zone mean air temperature, there is no difference between
                    # the model output and the forecasted temperate
                    forecasted_temp = y_hat_l[-1]
                    true_temp = ys
                    # if we predict the temperature difference, then the forecasted temperature is the sum of the
                    # previous temperature and the difference
                    if target_name == "Delta Mean Air Temperature(t+1)":
                        forecasted_temp += xs[temp_pos_idx]
                        true_temp += true_t_in
                    forecasted_temp_l.append(forecasted_temp)
                    true_temp_l.append(true_temp)
        y_true_np, y_hat_np, true_temp_np, forecasted_temp_np = map(
            lambda x: torch.cat(x).detach().cpu().numpy(), (y_true_l, y_hat_l, true_temp_l, forecasted_temp_l))
        metrics_y = c.compute_metrics(y_true_np, y_hat_np)
        metrics_temp = c.compute_metrics(true_temp_np, forecasted_temp_np)
        metrics_y = {k + "_24h": round(v, 3) for k, v in metrics_y.items()}
        metrics_temp = {k + "_24h": round(v, 3) for k, v in metrics_temp.items()}

    return metrics_y, metrics_temp


def metricsover24hours_rc(model, target_name, dl):
    """
    Compute the metrics over 24h for the RC model.
    :param model: the model
    :param target_name: the name of the target variable
    :param dl: a dataloader containing the data. The batch size should be such as to cover one day. And one batch
            should be such that, when iterating over it, one sample can be passed to the model directly.
    :return: two dictionaries of metrics. One contains the error metrics on the model output and the other on the
            forecasted temperature. The two dictionaries contain the same keys but the values are different
            if the target name is "Delta Mean Air Temperature(t+1)".
    """
    y_true_l, y_hat_l = [], []
    true_temp_l, forecasted_temp_l = [], []  # these are useful only for Delta Zone Mean Air Temperature(t+1) target name
    model.eval()
    with torch.no_grad():
        for i_b, (xb, yb) in enumerate(dl):
            if i_b % 4 == 0:
                forecasted_temp = xb[0]["t_in"]
                for xs, ys in zip(xb, yb):
                    true_t_in = xs["t_in"]
                    xs["t_in"] = forecasted_temp
                    y_hat_l.append(model(xs))
                    y_true_l.append(torch.reshape(ys, (1, -1)))
                    # if we predict directly the zone mean air temperature, there is no difference between
                    # the model output and the forecasted temperate
                    forecasted_temp = y_hat_l[-1]
                    true_temp = ys
                    # if we predict the temperature difference, then the forecasted temperature is the sum of the
                    # previous temperature and the difference
                    if target_name == "Delta Mean Air Temperature(t+1)":
                        forecasted_temp += xs["t_in"]
                        true_temp += true_t_in
                    forecasted_temp_l.append(torch.reshape(forecasted_temp, (1, -1)))
                    true_temp_l.append(torch.reshape(true_temp, (1, -1)))
        y_true_np, y_hat_np, true_temp_np, forecasted_temp_np = map(
            lambda x: torch.cat(x).detach().cpu().numpy(), (y_true_l, y_hat_l, true_temp_l, forecasted_temp_l))
        metrics_y = c.compute_metrics(y_true_np, y_hat_np)
        metrics_temp = c.compute_metrics(true_temp_np, forecasted_temp_np)
        metrics_y = {k + "_24h": round(v, 3) for k, v in metrics_y.items()}
        metrics_temp = {k + "_24h": round(v, 3) for k, v in metrics_temp.items()}

    return metrics_y, metrics_temp


def neuralnetworkall(data: pd.DataFrame, resp_name: list, hyperparam: dict, NN_path: str,
                     fig_path: str,
                     save_model: bool = False, save_fig: bool = False, seed: int = None, seed_db: int = 34,
                     steps_btw_batches: int = 96):
    """

    :param data:
    :param resp_name:
    :param NN_path:
    :param fig_path:
    :param save_model:
    :param save_fig:
    :param seed: is the seed for the NN training
    :param rnd_db: is the seed for the database randomization
    :return:
    """

    # define the NN
    class NeuralNetwork(nn.Module):
        def __init__(self, input_size, output_size, nb_hidden_layers, nb_neurons_per_layer, activation_function):
            super().__init__()
            self.flatten = nn.Flatten()
            layers = []  # Initialize a list to store the layers
            layers.append(nn.Linear(input_size, nb_neurons_per_layer))  # Add input_batch layer
            layers.append(activation_function)

            # Add hidden layers
            for _ in range(nb_hidden_layers - 1):
                layers.append(nn.Linear(nb_neurons_per_layer, nb_neurons_per_layer))
                layers.append(activation_function)

            # Add output layer
            layers.append(nn.Linear(nb_neurons_per_layer, output_size))

            self.full_stack = nn.Sequential(*layers)

        def forward(self, x):
            # must be removed ortherwise `write_onnx_model_with_bounds` does not work
            # x = self.flatten(x)  # Flatten return a copy of the input collapsed into one dimension
            prediction = self.full_stack(x)
            return prediction

    def MSEandLasso(nnet, outputs, targets, coeff=1e-3):
        """
        Compute the Mean Squared Error with L1 regularization.
        :param nnet: the neural network
        :param outputs: the predictions
        :param targets: the true values
        :param coeff: the coefficient for the L1 regularization
        :return: the loss
        """
        mse_loss = nn.MSELoss()
        mse_loss = mse_loss(outputs, targets)
        lasso_loss = 0
        for l in nnet.full_stack:
            if isinstance(l, nn.Linear):
                lasso_loss += l.weight.abs().sum()
        regularized_loss = mse_loss + coeff * lasso_loss
        return regularized_loss

    def train(nnet, loss_fct, reg_coeff, optimizer, dataloader):
        """
        Train the model over one epoch only.
        :param loss_fct:
        :param optimizer:
        :param num_epochs:
        :param dataloader:
        :return:
        """
        nnet.train()  # Set model to training mode
        for batch_n, (batch_inputs, batch_targets) in enumerate(dataloader):
            batch_inputs, batch_targets = batch_inputs.to(device), batch_targets.to(device)  # Load the data to the MPS

            # Forward pass
            outputs = nnet(batch_inputs)
            loss = loss_fct(nnet, outputs, batch_targets, reg_coeff)

            # Backward pass and optimization
            loss.backward()
            optimizer.step()
            optimizer.zero_grad()  # Zero the gradients

        # Print the loss for monitoring training progress
        print(f"Training loss: {loss.item():.4f} -- ", end='')
        return loss.item()

    def validation(nnet, loss_fct, reg_coeff, dataloader):
        num_batches = len(dataloader)
        nnet.eval()
        validation_loss = 0
        mae, mse, relmae, relmse, r_squared = ([], [], [], [], [])
        with torch.no_grad():
            for X, y in dataloader:
                X, y = X.to(device), y.to(device)
                pred = nnet(X)
                validation_loss += loss_fct(nnet, pred, y, coeff=reg_coeff).item()
                y = y.cpu().detach().numpy()  # numpy works only on the cpu so the data needs to be sent to the cpu...
                pred = pred.cpu().detach().numpy()  # ...before using any numpy function on them
                mae.append(np.mean(np.abs(y - pred)))
                mse.append(np.mean((y - pred) ** 2))
                relmae.append(np.mean(np.abs((y - pred) / y)))
                relmse.append(np.sqrt(np.mean(((y - pred) / y) ** 2)))
                r_squared.append(1 - sum((y - pred) ** 2) / sum((y - np.mean(y)) ** 2))

        validation_loss /= num_batches
        mae_avg, mse_avg, r_squared_avg = map(np.mean, (mae, mse, r_squared))
        std = np.std(y - pred)
        rel_std = np.std((y - pred) / y)
        print(f"Validation avg loss: {validation_loss:.4f}, MAE: {mae_avg:.4f}, MSE: {mse_avg:.4f}, "
              f"R\N{SUPERSCRIPT TWO}: {r_squared_avg:.4f}", end='\n')
        return validation_loss, {'MAE_val': mae_avg, 'MSE_val': mse_avg, 'relMAE_val': np.mean(relmae),
                                 'relMSE_val': np.mean(relmse), 'std': std, 'rstd': rel_std, 'R2_val': r_squared_avg}

    def launch_tensorboard(logdir, port=6007):
        tb = program.TensorBoard()
        tb.configure(argv=[None, '--logdir', logdir, '--port', str(port)])
        url = tb.launch()
        print(f'TensorBoard started at {url}')

    # HyperParameters
    sparse, activation, nb_layers, nb_neurons, batch_size, nb_epochs_max, patience_max, reg_coeff = hyperparam.values()
    seed = random.randint(0, 10000) if seed is None else seed

    # To make code reproducible
    seed, seed_worker, g = c.make_reproducible(seed)

    # Define the path for the folder which will contain all the outputs of this specific model
    current_time = datetime.now().strftime("%Y-%m-%d %Hh%Mm%Ss")
    folder_path = os.path.join(NN_path, f"NotSparse/NN_{nb_layers}layers_{nb_neurons}neurons/{current_time}")
    figure_path = os.path.join(fig_path, f"NotSparse/NN_{nb_layers}layers_{nb_neurons}neurons")
    createdir(folder_path, up=2)
    createdir(figure_path, up=1)

    # Train the model on GPU if available (on devices with MacOS, GPU = MPS)
    device = "cpu"  # c.get_available_device() or "cpu"
    print(f"\nUsing {device} device")

    # Normalize input_batch
    data_norm, scalers = c.normalizedata(data, resp_name)
    db_norm = Database(data_norm, resp_name, seed_db, shuffle=True)

    # Put data into a TensorDataset
    training_set = TensorDataset(torch.tensor(db_norm.X_train.values, dtype=torch.float32),
                                 torch.tensor(db_norm.y_train_true.values, dtype=torch.float32))
    validation_set = TensorDataset(torch.tensor(db_norm.X_test.values, dtype=torch.float32),
                                   torch.tensor(db_norm.y_test_true.values, dtype=torch.float32))

    # Create a DataLoader to efficiently load the data during training
    training_dataloader = DataLoader(training_set, batch_size=batch_size, shuffle=True, worker_init_fn=seed_worker,
                                     generator=g)
    validation_dataloader = DataLoader(validation_set, batch_size=batch_size * 2, shuffle=True,
                                       worker_init_fn=seed_worker, generator=g)

    # Create the model
    nnet = NeuralNetwork(input_size=len(db_norm.X.columns), output_size=len(db_norm.y_true.columns),
                         nb_hidden_layers=nb_layers, nb_neurons_per_layer=nb_neurons,
                         activation_function=activation).to(device)
    print(nnet)

    # Define the loss function and the optimizer
    loss_fct = MSEandLasso  # Mean Squared Error with L1 regularization
    reg_coeff = hyperparam['regularization_coeff']
    # Define the optimizer, I advise Adam (quicker)
    # Set weight_decay > 0  for ridge (L2) regularization
    optimizer = optim.Adam(nnet.parameters(), lr=0.001, weight_decay=0)
    # optimizer = optim.SGD(nnet.parameters(), lr=0.001, momentum=0.9)

    # Train and test the model
    tensorboard_log_dir = os.path.join(c.path_up(folder_path, 1), f"Tensorboard/{current_time}")
    writer = SummaryWriter(log_dir=tensorboard_log_dir)
    best_loss = np.inf
    patience, nb_epochs = 0, 0
    start_time = time.time()
    for epoch in range(nb_epochs_max):
        if patience >= patience_max:
            print(f'\nEarly stopping at epoch {epoch + 1}')
            break
        # Print the loss for monitoring training progress
        print(f"Epoch [{epoch + 1}/{nb_epochs_max}]: ", end='')
        train_loss = train(nnet, loss_fct, reg_coeff, optimizer, training_dataloader)
        validation_loss, metrics_val = validation(nnet, loss_fct, reg_coeff, validation_dataloader)
        patience += 1
        nb_epochs += 1  # to be saved in summary file
        if validation_loss < best_loss:
            patience = 0
            best_loss = validation_loss
            best_metrics_val = metrics_val
            best_weights = nnet.state_dict()
        # Log the running loss averaged per batch
        writer.add_scalars('Training vs. Validation Loss', {'Training': train_loss, 'Validation': validation_loss},
                           epoch + 1)
    training_time = time.time() - start_time  # to be saved in summary file
    nnet.load_state_dict(best_weights)  # Load the best weights
    writer.flush()  # wave the content of the writer buffer into the disk

    # Grab a single batch of the training set
    dataiter = iter(training_dataloader)
    input_batch, _ = next(dataiter)

    # add_graph() traces the sample input_batch through your model, and renders it as a graph.
    writer.add_graph(nnet, input_batch.to(device), verbose=False)
    writer.flush()
    writer.close()

    # evaluate over 24h forecast
    db = Database(data_norm, resp_name, shuffle=False)
    test_24h_ds = TensorDataset(torch.tensor(db.X.values.astype(float), dtype=torch.float32),
                                  torch.tensor(db.y_true.values.astype(float), dtype=torch.float32))
    def sample_to_device(x, y):
        return x.to(device), y.to(device)
    test24h_dl = WrappedDataLoader(DataLoader(test_24h_ds, batch_size=24, shuffle=False, worker_init_fn=seed_worker,
                                              generator=g, drop_last=True), sample_to_device)
    temp_col = [i for i, b in enumerate(db.X.columns.str.contains("Mean Air Temperature")) if b]
    _, metrics24h = metricsover24hours_nn(nnet, resp_name[0].split(",")[0], test24h_dl, temp_col)

    # # Plot 4 days (picked randomly)
    # for batch_num in [30]:
    #     tempforecast(y_true_batch[batch_num], y_hat_batch[batch_num], batch_num, save_fig=save_fig,
    #                  folderpath=figure_path,
    #                  filename=f"24hForecast_Batch{batch_num}")

    if save_model:
        joblib.dump(scalers, os.path.join(folder_path, 'scaler.pkl'))  # Save the scaler into a pickle file
        # Save the NN as ONNX file for the OMLT package to reformulate it into pyomo
        savetoonnx(folder_path, nnet, to_onnx, db_norm, input_batch, scalers['input'])

        # Save the hyperparameters of the model and its performance in validation and 24h-ahead forecast to summary file
        hyperparameters = {"time": current_time, "sparse": sparse, "target": resp_name[0].split(",")[0],
                           "activation": activation.__str__(), "nb_layers": nb_layers, "nb_neurons": nb_neurons,
                           "nb_inputs": len(db_norm.X.columns), "seed": seed, "batch_size": batch_size,
                           "nb_epochs": nb_epochs,
                           "training_time": np.round(training_time, decimals=3)} | best_metrics_val
        # append the MAE, MSE and R2 computed over 24h forecasts
        hyperparameters.update({k: np.round(np.mean(v), decimals=4) for k, v in metrics24h.items()})
        savetosummary(hyperparameters, os.path.join(c.path_up(folder_path, 3), 'TrainingSummary_NN2.xlsx'), sparse)

    # Launch TensorBoard
    # launch_tensorboard(tensorboard_log_dir, port=6007)

    return data_norm, db_norm, scalers


def to_onnx(filepath, model, input_sample):
    """
    Save pytorch model to onnx format.
    :param filepath:
    :param model:
    :param input_sample: a tuple containing the input data (a tensor)
    :return:
    """
    torch.onnx.export(
        model,
        input_sample,
        filepath,
        input_names=['input'],
        output_names=['output'],
        dynamic_axes={'input': {0: 'batch_size'}, 'output': {0: 'batch_size'}}
    )


def treetoonnx(filepath, model, input_sample):
    # Convert into ONNX format
    from skl2onnx import convert_sklearn
    from skl2onnx.common.data_types import FloatTensorType
    initial_type = [('input', FloatTensorType([None, input_sample.shape[0]]))]
    final_type = [('output', FloatTensorType([None, 1]))]
    onx = convert_sklearn(model, initial_types=initial_type, final_types=final_type)
    with open(filepath, "wb") as f:
        f.write(onx.SerializeToString())


def savetoonnx(folderpath, model, modeltoonnx, db: Database, input_sample, input_scaler):
    """
    Function to save the NN model and the input bounds to ONNX format (with the OMLT package)
    :param db: the Database object which contains both train and test data NORMALIZED
    :param model: the nn model
    :param input_sample: a batch of the input data (normalized)
    :param filepath: where to save the onnx model (with the input bounds)
    :return:
    """
    lb = db.X.min(axis=0).values
    ub = db.X.max(axis=0).values
    n_features = len(db.X.columns)
    # if the Unitary System Total Heating Rate [W] is included in the input, set the lower bound to 0 to cancel
    # ventilation consumption
    for j, f in enumerate(db.X.columns):
        if "Unitary System Electricity Rate [W]" in f:
            # set lower bound to 0 if it is positive (eliminate the constant consumption of the ventilation)
            if lb[j] > 0:
                lb[j] = input_scaler.transform(np.zeros((1, n_features)))[0, j]
        elif "Electricity:Facility [Wh]" in f:
            # set lower bound to 0 (eliminate the constant consumption of the ventilation)
            if lb[j] > 0:
                lb[j] = input_scaler.transform(np.zeros((1, n_features)))[0, j]
            # set upper bound to += 5000W to account for consumption peaks in real life
            raise NotImplementedError("Pay attention to the goal here and check the bounds")
            ub[j] += 14114
        elif "Hour" in f:
            # set upper bound to 24 in case there are hour fraction (e.g., 23.5 for 23h30)
            tmp = np.zeros((1, n_features))
            tmp[0, j] = 24
            ub[j] = input_scaler.transform(tmp)[0, j]
    input_bounds = [(l, u) for l, u in zip(lb, ub)]
    # Save the model as ONNX file
    modeltoonnx(os.path.join(folderpath, "model.onnx"), model, input_sample)
    write_input_bounds(os.path.join(folderpath, "Bounds.json"), input_bounds)  # write the bounds with OMLT package
    # /!\ do not use write_onnx_model_with_bounds() because it complexifies the reformulation in pyomo


def savetosummary(training_summary: dict, excel_path: str, sparse: bool = False):
    """
    Save the training summary to an excel file. The training summary usually contains the hyperparameters of the model
    and its performances during training (time) and test (metrics).
    If the excel file does not exist, it is created.
    :param training_summary: a dictionary containing the hyperparameters and the performances of the model
    :param excel_path: the path to the excel file
    :param sparse: whether the model is sparse or not
    :return: None
    """
    sheetname = "Sparse" if sparse else "NotSparse"

    # if the file does not exist, create the file, then create the sheet, then write the header and the summary
    # else, if the file exists, open it, then if the sheet doesn't, create it, then write the header and the summary
    if not os.path.exists(excel_path):  # if the file does not exist
        workbook = openpyxl.Workbook()  # create it
        sheet = workbook.active  # get the active sheet
        sheet.title = sheetname  # rename it
        sheet.append(list(training_summary.keys()))  # write the header
    else:  # if the file exists, open it
        workbook = openpyxl.load_workbook(excel_path)
        if sheetname not in workbook.sheetnames:  # if the sheet does not exist
            sheet = workbook.create_sheet(sheetname)  # create it
            sheet.append(list(training_summary.keys()))  # write the header
        else:  # if the sheet exists, select it
            sheet = workbook[sheetname]

    # Save a summary of the results
    sheet.append(list(training_summary.values()))
    workbook.save(excel_path)

    print(training_summary)
